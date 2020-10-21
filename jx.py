from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import traceback
import os
from selenium.webdriver.chrome.options import Options
import openpyxl
import traceback
import re
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.select import Select
import pyautogui #鼠标点击
import datetime
from selenium.webdriver.common.keys import Keys
from emailSMTP import emailtest #引用邮件
#启动Chrome;运行cmd的命令
def start_chrome():
	chrome_exe = 'chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile'
	os.popen(chrome_exe)
	time.sleep(3)

def dr():
	chrome_options = Options()
	chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
	chrome_driver = "C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe"
	driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)
	return driver

def login_url():
	GENERATION_URL='https://hc.ikandy.cn/login'
	TEST_URL = 'https://jx-h5-test.ikandy.cn:4001/login'
	driver.get(GENERATION_URL)
	print('before search======')
	#打印当前页面title
	now_title=driver.title
	#打印当前也的url
	now_url=driver.current_url

#登录
def login_account():
	driver.find_element_by_id('loginName').send_keys('zyyadmin1')
	driver.find_element_by_id('password').send_keys('yy880914')
	time.sleep(1)
	dengl=driver.find_elements_by_class_name('txt-login-form-button')
	print(len(dengl))
	dengl[0].send_keys(Keys.ENTER)
	#ActionChains(driver).click_and_hold().perform()
	time.sleep(2)
	#异常判断，并把异常保存在本地
	try:
		driver.find_element_by_xpath('//*[@id="root"]/div/section/div[2]/div/section/aside/div[1]/ul/li[2]/div/span').click()
	except :
		errorfile=open('erroe.txt','w+')
		errorfile.write(traceback.format_exc())
		errorfile.close()
	finally:
		print(time.ctime())
		print('+++++++++++++++++')
	time.sleep(2)

file_name = 'CaseV2-20200308.xlsx'
filename = ''.join(['F:\MyDownloads\\',file_name])
save_filename = ''.join(['F:\MyDownloads\A',file_name])
#读取表格的值
def read_xlsx():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	print(sheet.max_row)
	currt_id = sheet['A2'].value
	sheet['B1'] = '坐席'
	for currt_id_raw in range(2,sheet.max_row+1):
		produceName = sheet.cell(row=currt_id_raw,column=1).value
		clear_find_name()
		find_name(produceName)
		search_button()
		get_name_id()
		sheet['B'+str(currt_id_raw)] = get_name_id()
		clear_find_name()
	wb.save(save_filename)

	print(currt_id)
	time.sleep(1)

def search_account():
	#进入工单
	sum_ele = driver.find_elements_by_class_name('ant-menu-submenu-title')
	sum_ele_button = driver.find_elements_by_class_name('ant-menu-submenu-arrow')
	print(len(sum_ele))
	print(len(sum_ele_button))
#	sum_ele_button[0].send_keys(Keys.ENTER)
	sum_ele_button[0].click()
	sum_elements = driver.find_elements_by_class_name('ant-menu-item')
	print(len(sum_elements))
#根据来源填入查询姓名身份证的框
def find_name(test):
	search_name = driver.find_element_by_id('customerName')
	search_name.send_keys(test)
#打印出坐席的名字
def get_name_id():
	time.sleep(2)
	names = driver.find_elements_by_class_name('ant-table-row')
	namess = driver.find_element_by_xpath('//*[@id="root"]/div/section/div[2]/div[2]/section/section/div/div/main/div[2]/div[3]/div/div/div/div/div/div[2]/div/div[1]/table/tbody/tr[1]/td[14]')
	return namess.text

def search_button():
	driver.find_element_by_class_name('ant-btn').send_keys(Keys.ENTER)
def clear_find_name():
	search_name = driver.find_element_by_id('customerName')
	search_name.clear()

#设置账户管理
def zh_gl():
	wb = openpyxl.load_workbook('20200310reportsave.xlsx')
	sheet = wb.get_active_sheet()
	for i in range(1,2):
		name = sheet.cell(row=i,column=1).value
		name_input = driver.find_element_by_class_name('ant-input')
		name_input.clear()
		name_input.send_keys('Yu zi')
		time.sleep(2)
		driver.find_element_by_class_name('ant-input-search-button').send_keys(Keys.ENTER)
		time.sleep(2)
		driver.find_element_by_xpath('//*[@id="root"]/div/section/div[2]/div/section/section/div/div/main/div[2]/div[3]/div/div/div/div/div/div/div[1]/table/tbody/tr[1]/td[9]/span/a[2]').send_keys(Keys.ENTER)
		time.sleep(2)
		driver.find_elements_by_class_name('ant-btn-primary')[2].send_keys(Keys.ENTER)
		time.sleep(2)

#截图
def current_photo():
	data=time.strftime('%m%d-%H%M',time.localtime(time.time()))
	a=data+'.jpg'
	driver.get_screenshot_as_file(a)

if __name__ == '__main__':
	start_chrome()
	driver=dr()

	print('开始')
	login_url()
	login_account()
	time.sleep(10)
	zh_gl()
#	driver.maximize_window()
#	read_xlsx()
#	search_account()
#	emailtest.run_send_mail(save_filename)