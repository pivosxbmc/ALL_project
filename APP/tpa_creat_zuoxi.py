from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import traceback
import os
from selenium.webdriver.chrome.options import Options
import openpyxl
import traceback
import re
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.select import Select
import pyautogui #鼠标点击
import datetime
from selenium.webdriver.common.keys import Keys
import requests
import pyperclip
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

#启动Chrome;运行cmd的命令
def start_chrome():
	chrome_exe = 'chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile'
	os.popen(chrome_exe)
	time.sleep(3)

def dr():
	chrome_options = Options()
	chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
	chrome_driver = "C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe"
	#chrome_driver = "C:\\Users\\10944\\AppData\\Local\\Google\\Chrome\\Application\\chromedriver.exe"
	driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)
	return driver

def login_url():
    a = driver.current_url
    if a=='https://tpaprod.ikandy.cn/txtechnology/login':
        pass
    else:
        driver.get('https://tpaprod.ikandy.cn')
        #处理网页弹框
        time.sleep(3)
        try:
        	driver.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[1]/div/div/div[2]/button').click()
        except Exception as e:
        	pass


#登录
def login_account(username,password):
	driver.find_elements_by_class_name('login-input')[0].send_keys(username)
	driver.find_elements_by_class_name('login-input')[1].send_keys(password)
	time.sleep(1)
	driver.find_element_by_class_name('login-btn').click()
	time.sleep(5)
	#driver.find_elements_by_class_name('offline')[0].click()




filename = '109449.xlsx'
#读取表格的值
def read_xlsx():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	for currt_id_raw in range(1,sheet.max_row+1):
		#sheet.max_row+1
		produceName = sheet.cell(row=currt_id_raw,column=1).value	
		driver.find_element_by_class_name('ant-input').send_keys('ASHH46037120EAAAY514')
		driver.find_element_by_class_name('ant-input-suffix').click()
		time.sleep(2)
		driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div/div[2]/div/div[2]/div[9]/div/div/div/div/div/div/div/div/table/tbody/tr/td[9]/div/span[3]').click()
		time.sleep(1)
		driver.find_element_by_class_name('ant-input').clear()
	time.sleep(1)

def beizhu():
	driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[2]/div/div[2]/div[1]/div/div/div/div/button[4]').click()
	time.sleep(2)
	beizhu_data = driver.find_element_by_class_name('ant-timeline-item').text
	print(beizhu_data)
	driver.find_element_by_class_name('ant-modal-close-x').click()
	time.sleep(1)


 
def read_xlsx_data(list_data_int,save_xlsx_name):
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	case_data = 3
	for currt_id_raw in range(list_data_int,sheet.max_row+1):
		#sheet.max_row+1
		time.sleep(2)
		driver.find_element_by_class_name('ant-input').clear()
		produceName = sheet.cell(row=currt_id_raw,column=1).value	
		driver.find_element_by_class_name('ant-input').send_keys(produceName)
		driver.find_element_by_class_name('ant-input-suffix').click()
		time.sleep(1)
		report_table_button = (By.CLASS_NAME,'report-table-list-operate')
		try:
			WebDriverWait(driver,10,1).until(EC.presence_of_element_located(report_table_button),'没有这个')
			driver.find_element_by_class_name('report-table-list-operate').click()
		except Exception as e:
			print(produceName+'='+str(currt_id_raw))
			wb.save(save_xlsx_name)
			raise e
		time.sleep(3)

		'''
		staus_text=driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[1]/span[2]').text
		if '多次报案'in staus_text:
			print(produceName)
			time.sleep(1)
		else:
			time.sleep(1)
		

		kind_data = driver.find_element_by_id('riskName').get_attribute('value')
		print(kind_data)
		if '雇主' in kind_data:
			print('滑动')
			#driver.execute_script("window.scrollTo(0,8500)")
			#ant-select-selection__rendered accidentPlace

			money = driver.find_element_by_id('claimAmount').get_attribute('value')
			if money:
				print('有索赔金额')
				driver.find_element_by_id('lossAssessmentAmount').send_keys(int(int(money)*1.5))
			else:
				driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[2]/div/div[2]/div[1]/div/div/div/div/button[4]').click()
				time.sleep(2)
				beizhu_data = driver.find_element_by_class_name('ant-timeline').text
				print(beizhu_data)
				if beizhu_data:
					pass
				else:
					beizhu_data='报警轻伤回访'
				regexd=re.compile(r'报警(.+)回访')
				xd=regexd.search(beizhu_data)
				sheet.cell(row=currt_id_raw,column=11).value = xd.group(0)
				print(xd.group(0))
				driver.find_element_by_class_name('ant-modal-close-x').click()
				time.sleep(1)
				if '骨折' in beizhu_data:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(30000)
				elif '机动车受损' in beizhu_data:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(10000)
				elif '非机动车 受损' in beizhu_data:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(1500)
				else:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(2000)
				time.sleep(2)
			ele2 = driver.find_elements_by_class_name('ant-select-selection')[2]
			text_chuxian = ele2.text
			if text_chuxian:
				pass
			else:
				driver.find_element_by_id("accidentPlace").send_keys(Keys.TAB)
				time.sleep(2)
				ele2 = driver.find_elements_by_class_name('ant-select-selection')[2]
				ele2.click()
				driver.switch_to.active_element.send_keys(Keys.ENTER)
				time.sleep(2)

		else:
			time.sleep(2)
			#driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
			money = driver.find_element_by_id('claimAmount').get_attribute('value')
			if money:
				print('有索赔金额')
				driver.find_element_by_id('lossAssessmentAmount').send_keys(int(int(money)*1.5))
			else:
				driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[2]/div/div[2]/div[1]/div/div/div/div/button[4]').click()
				time.sleep(2)
				beizhu_data = driver.find_element_by_class_name('ant-timeline').text
				print(beizhu_data)
				if beizhu_data:
					pass
				else:
					beizhu_data='报警轻伤回访'
				regexd=re.compile(r'报警(.+)回访')
				xd=regexd.search(beizhu_data)
				sheet.cell(row=currt_id_raw,column=11).value = xd.group(0)
				print(xd.group(0))
				driver.find_element_by_class_name('ant-modal-close-x').click()
				time.sleep(1)
				if '骨折' in beizhu_data:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(30000)
				elif '机动车受损' in beizhu_data:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(10000)
				elif '非机动车' in beizhu_data:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(1500)
				else:
					driver.find_element_by_id('lossAssessmentAmount').send_keys(2000)
				time.sleep(1)
			print('滑动')
			ele2 = driver.find_elements_by_class_name('ant-select-selection')[2]
			text_chuxian = ele2.text
			if text_chuxian:
				print('已经有值：'+str(produceName))

			else:
				driver.find_element_by_id("accidentPlace").send_keys(Keys.TAB)
				time.sleep(2)
				ele2.click()
				#driver.switch_to.active_element.send_keys(Keys.DOWN)
				driver.switch_to.active_element.send_keys(Keys.ENTER)
				time.sleep(2)
				ele=driver.find_elements_by_class_name('ant-select-selection')[3]
				ele.click()
				driver.switch_to.active_element.send_keys(Keys.DOWN*3)
				driver.switch_to.active_element.send_keys(Keys.ENTER)

		driver.find_elements_by_class_name('ant-btn-sm')[0].click()
		time.sleep(3)
		'''

		#照片取证
		time.sleep(2)
		photo_tab_element = (By.CLASS_NAME,'ant-tabs-tab')
		WebDriverWait(driver,10,1).until(EC.presence_of_element_located(photo_tab_element),'没有这个')
		driver.find_elements_by_class_name('ant-tabs-tab')[1].click()
		time.sleep(1)
		try:
			data_photo=driver.find_elements_by_class_name('ant-tabs-nav-animated')[1].text
		except Exception as e:
			print(currt_id_raw)
			wb.save(save_xlsx_name)
			raise e
		regexx=re.compile(r'(\d+)')
		xdc=regexx.findall(data_photo)
		sum=int((xdc)[0])+int((xdc)[1])+int((xdc)[2])+int((xdc)[3])+int((xdc)[4])+int((xdc)[5])
		sheet.cell(row=currt_id_raw,column=11).value = sum
		time.sleep(1)
		'''
		#进入工单详情
		driver.find_elements_by_class_name('ant-tabs-tab')[0].click()		
		#备注
		driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div[2]/div/div[2]/div[1]/div/div/div/div/button[4]').click()
		time.sleep(2)
		beizhu_data = driver.find_element_by_class_name('ant-timeline-item').text
		print(beizhu_data)
		driver.find_element_by_class_name('ant-modal-close-x').click()
		time.sleep(1)
		'''


		#退回
		if sum <= 5:
			case_data +=1
			print(produceName)
			print(case_data)
			data_tuihui='【太平洋保险】尊敬的客户，您的理赔申请缺少必要材料，请进入“太保代客服助手”微信小程序首页点击“查看索赔资料”，按投保险种，查看并下载相应理赔资料清单，如已提交相应资料请略过，如未提交，请尽快上传，以免影响后续理赔。'
			driver.find_elements_by_class_name('ant-btn-primary')[5].click()
			time.sleep(2)
			try:
				sendtext = driver.find_element_by_id('input_leaveMsg')
			except Exception as e:
				wb.save(save_xlsx_name)
				driver.find_elements_by_class_name('ant-btn-primary')[6].click()
				#driver.find_elements_by_class_name('ant-btn')[13].click()
				time.sleep(1)
				sendtext = driver.find_element_by_id('input_leaveMsg')
			sendtext.click()
			sendtext.send_keys(data_tuihui)
			time.sleep(1)
			'''
			pyautogui.click(482,521)
			pyperclip.copy(data_tuihui)
			pyautogui.hotkey('ctrl','v')
			'''
			driver.find_element_by_xpath('//*[@id="addLeaveMsg"]').click()
			driver.find_element_by_class_name('ant-modal-close-x').click()
			time.sleep(1)
			driver.find_element_by_class_name('back_to_list_span').click()
			time.sleep(2)
		else:
			driver.find_element_by_class_name('back_to_list_span').click()
			time.sleep(1)

		#driver.find_element_by_class_name('back_to_list_span').click()
	time.sleep(1)
	wb.save(save_xlsx_name)
	print('完成')


#截图
def current_photo():
	data=time.strftime('%m%d-%H%M',time.localtime(time.time()))
	a=data+'.jpg'
	driver.get_screenshot_as_file(a)


def posxy():
	x,y=pyautogui.position() #获取当前鼠标的位置
	print(x,y)
#TPA创建账号
def choice_account(loginName,fullName,cellphone,password='123qwer'):
	time.sleep(2)
	#driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div[1]/div[4]/input').clear()
	#driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div[1]/div[4]/input').send_keys(loginNamesearch)
	#driver.find_element_by_class_name('btn_search').click()
	driver.find_element_by_class_name('btn_create').click()
	time.sleep(1)
	#driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div[2]/div[1]/div/div/div/div/div/table/tbody/tr/td[9]/div/span/a').click()
	#pyautogui.click(425,182) #点击创建按钮
	#driver.find_element_by_class_name('btn_create').click()
	time.sleep(1)
	driver.find_element_by_id('loginName').clear()
	driver.find_element_by_id('loginName').send_keys(loginName) #账号
	driver.find_element_by_id('fullName').send_keys(fullName)  #用户名
	driver.find_element_by_id('cellphone').send_keys(cellphone)
	choice_role = driver.find_elements_by_class_name('ant-select-selection__placeholder')
	choice_role[5].click()
	driver.switch_to.active_element.send_keys(Keys.DOWN)
	driver.switch_to.active_element.send_keys(Keys.ENTER)
	element_sum_supervisePhone = driver.find_element_by_id('supervisePhone').send_keys(cellphone)
	element_sum_password = driver.find_element_by_id('password').clear()
	element_sum_password = driver.find_element_by_id('password').send_keys(password)
	element_sum_repassword = driver.find_element_by_id('repassword').send_keys(password)
	driver.find_elements_by_class_name('ant-checkbox-input')[-2].click()
	driver.find_elements_by_class_name('ant-checkbox-input')[-3].click()
	time.sleep(1)
	driver.find_elements_by_class_name('ant-checkbox-input')[-4].click()
	time.sleep(1)
	driver.find_elements_by_class_name('ant-checkbox-input')[-5].click()
	#创建确定按钮
	driver.find_element_by_class_name('ant-btn-primary').click()
	time.sleep(2)
	
def read_xlsx_zuoxi(filename):
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	for currt_id_raw in range(6,11): #,sheet.max_row+1):
		#sheet.max_row+1
		#姓名、账号、手机号，密码默认123qwe
		user_name = sheet.cell(row=currt_id_raw,column=2).value
		user_loginname = sheet.cell(row=currt_id_raw,column=1).value
		print(user_loginname)
		user_phone = sheet.cell(row=currt_id_raw,column=3).value
		choice_account(user_loginname,user_name,user_phone)


if __name__ == '__main__':
	#pyautogui.click(482,521)
	start_chrome()
	driver=dr()
	login_url()
	login_account('cbjg-admin','123qwer')
	time.sleep(10)
	#driver.find_elements_by_class_name('left-info')[3].click()
	read_xlsx_zuoxi('testname.xlsx')
	#read_xlsx_data(2,'18266248305.xlsx')
	#posxy()


