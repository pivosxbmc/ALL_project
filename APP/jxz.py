from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import traceback
import os
from selenium.webdriver.chrome.options import Options
import openpyxl
import traceback
import re
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.select import Select
import pyautogui #鼠标点击
import datetime
from selenium.webdriver.common.keys import Keys
import requests
#启动Chrome;运行cmd的命令
def start_chrome():
	chrome_exe = 'chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile'
	os.popen(chrome_exe)
	time.sleep(3)

def dr():
	chrome_options = Options()
	chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
	chrome_driver = "C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe"
	driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)
	return driver

def login_url(tpa_url):
    driver.get(tpa_url)



#登录
def login_account():
	driver.find_elements_by_xpath('//*[@id="center"]/form/table/tbody/tr[3]/td[2]/input').send_keys('admin')
	driver.find_element_by_xpath('//*[@id="center"]/form/table/tbody/tr[4]/td[2]/input[1]').send_keys('admin')
	time.sleep(1)
	driver.find_element_by_id('idConfirm').click()
	time.sleep(1)

def account(DisplayName,AuthName,UserName,editPassword):
	driver.find_element_by_id('account').click()
	driver.find_element_by_id('DisplayName').send_keys(DisplayName)
	driver.find_element_by_id('AuthName').send_keys(AuthName)
	driver.find_element_by_id('UserName').send_keys(UserName)
	driver.find_element_by_id('editPassword').send_keys(editPassword)
	driver.find_element_by_id('server1').send_keys('106.14.99.6')
	driver.find_element_by_id('port1').send_keys('5060')

def read_xlsx_tap():	
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	for i in range(a,b):
		a_name = sheet.cell(row=currt_id_raw,column=1).value
		b_data = 




filename = '109449.xlsx'
#读取表格的值
def read_xlsx():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	for currt_id_raw in range(283,418):
		#sheet.max_row+1
		produceName = sheet.cell(row=currt_id_raw,column=1).value	
		driver.find_element_by_class_name('ant-input').send_keys(produceName)
		driver.find_element_by_class_name('ant-input-suffix').click()
		time.sleep(2)
		driver.find_element_by_xpath('//*[@id="root"]/div/div[1]/div[2]/div/div[2]/div/div/div[2]/div/div/div/div[2]/div/div[2]/div[9]/div/div/div/div/div/div/div/div/table/tbody/tr/td[9]/div/span[3]').click()
		time.sleep(1)
		driver.find_element_by_class_name('ant-input').clear()
	time.sleep(1)


def get_url(file_url):
	time.sleep(1)
	url_test = ''.join(['https://livecenter-prod.ikandy.cn/api/recordVideos?taskId=',file_url,'&channel=aviva_org'])
	driver.get(url_test)
def get_url_txt():
	time.sleep(1)
	txturl = driver.find_element_by_xpath('/html/body/pre')
	#print(txturl.text)
	regexx=re.compile(r'http.*mp4')
	xd=regexx.findall(txturl.text)
	time.sleep(1)
	return str(xd)


def search_account():
	#进入工单
	sum_ele = driver.find_elements_by_class_name('ant-menu-submenu-title')
	sum_ele_button = driver.find_elements_by_class_name('ant-menu-submenu-arrow')
	print(len(sum_ele))
	print(len(sum_ele_button))
#	sum_ele_button[0].send_keys(Keys.ENTER)
	sum_ele_button[0].click()
	sum_elements = driver.find_elements_by_class_name('ant-menu-item')
	print(len(sum_elements))
#根据来源填入查询姓名身份证的框
def find_name(test):
	search_name = driver.find_element_by_id('customerName')
	search_name.send_keys(test)
#打印出坐席的名字
def get_name_id():
	time.sleep(2)
	names = driver.find_elements_by_class_name('ant-table-row')
	namess = driver.find_element_by_xpath('//*[@id="root"]/div/section/div[2]/div[2]/section/section/div/div/main/div[2]/div[3]/div/div/div/div/div/div[2]/div/div[1]/table/tbody/tr[1]/td[14]')
	return namess.text

def search_button():
	driver.find_element_by_class_name('ant-btn').send_keys(Keys.ENTER)
def clear_find_name():
	search_name = driver.find_element_by_id('customerName')
	search_name.clear()


#截图
def current_photo():
	data=time.strftime('%m%d-%H%M',time.localtime(time.time()))
	a=data+'.jpg'
	driver.get_screenshot_as_file(a)




#访问接口获取数据
def guoren_requests():
	r = requests.get(b_data,auth=('ymf-admin','123456'))
	print(r.text)
	regexx=re.compile(r'count":(\d+),"')
	xd=regexx.findall(r.text)
	B_data = int(xd[0])

	rc = requests.get(c_data)
	regexx=re.compile(r'count":(\d+),"')
	xdc=regexx.findall(rc.text)
	C_data = int(xdc[0])
	return [B_data,C_data]



def tb_data(url):
	url_address = requests.get(url)
	regexx=re.compile(r'count":(\d+),"')
	xdc=regexx.findall(url_address.text)
	sum_data = int(xdc[0])
	return sum_data


#读取URL
def read_xlsxx():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	print(sheet.max_row)
	currt_id = sheet['A5'].value
	sheet['C1'] = 'URL'
	for currt_id_raw in range(5,50):
		#sheet.max_row+1
		produceName = sheet.cell(row=currt_id_raw,column=1).value
		sheet['C'+str(currt_id_raw)] = get_requests(produceName)
	wb.save(save_filename)

	print(currt_id)
	time.sleep(1)

def read_xlsx_dzh():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	filename = 'D:\jx\wenjian\周报数据汇总.xlsx'
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	sheet['G3'] = guoren_requests()[0]
	sheet['F3'] = guoren_requests()[1]
	sheet['D6'] = tb_data(tb_URL)
	sheet['D9'] = huanhe_data(hh_URL)
	wb.save(filename)
	time.sleep(1)

#读取时间表
def time_sum():
	filename = r'C:\Users\DELL\Documents\WeChat Files\sixin2010123\FileStorage\20200320report.xlsx'
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	print(sheet.max_row+1)
	for rowNum in range(2,sheet.max_row+1):
		produceName = sheet.cell(row=rowNum,column=1).value
		xin = produceName[3:5]
		sheet.cell(row=rowNum,column=10).value = xin
		xin2 = produceName[-2:]
		sheet.cell(row=rowNum,column=11).value = xin2
	wb.save(filename)


if __name__ == '__main__':

	start_chrome()
	driver=dr()
	login_url()
	login_account()
	time.sleep(5)
	read_xlsx()
