# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import traceback
import re
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.select import Select
import pyautogui #鼠标点击
import datetime
import os
from selenium.webdriver.common.keys import Keys
import pyperclip #复制粘贴
def dr():
	driver = webdriver.Chrome()
	return driver
def current_url():
	a=driver.current_url
	if a=='https://yds-dzh-prod.ikandy.cn//txtechnology/login':
		#driver.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[1]/div/div/div[2]/button').click()
		pass
	else:
		driver.get('https://yds-dzh-prod.ikandy.cn/')
		#处理网页弹框
		time.sleep(3)
		driver.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[1]/div/div/div[2]/button').click()
def login():
	loging_inputs=driver.find_elements_by_class_name('form-control')
	loging_inputs[0].clear()
	loging_inputs[0].send_keys('qdrb')
	loging_inputs[1].clear()
	loging_inputs[1].send_keys('123456')
	time.sleep(15)
	driver.find_element_by_class_name('ant-btn').click()
	time.sleep(2)
def setting_button():
	buttonsetting = driver.find_elements_by_class_name('left-info')
	buttonsetting[2].click()
def choice_account(loginName,fullName,cellphone,password):
	'''创建多租户账号的模板 适用机构管理员'''
	time.sleep(1)
	#点击创建按钮
	driver.find_element_by_class_name('btn_create').click()
	time.sleep(2)
	element_sum_loginName = driver.find_element_by_id('loginName').send_keys(loginName)
	element_sum_fullname = driver.find_element_by_id('fullName').send_keys(fullName)
	element_sum_cellphone = driver.find_element_by_id('cellphone').send_keys(cellphone)
	choice_role = driver.find_elements_by_class_name('ant-select-enabled')
	choice_role[1].click()
	driver.switch_to.active_element.send_keys(Keys.DOWN)
	driver.switch_to.active_element.send_keys(Keys.ENTER)
	element_sum_supervisePhone = driver.find_element_by_id('supervisePhone').send_keys('123456')
	element_sum_password = driver.find_element_by_id('password').send_keys(password)
	element_sum_repassword = driver.find_element_by_id('repassword').send_keys(password)
	element_sum_permission = driver.find_element_by_id('permission').click()
	#element_sum_pmsnAssign = driver.find_element_by_id('pmsnAssign').click()
	#创建确定按钮
	driver.find_element_by_class_name('ant-btn-primary').click()
	time.sleep(2)

def create_jg(orgname,EnglishName,address):
	driver.find_element_by_class_name('btn_create').click()
	time.sleep(1)
	driver.find_element_by_id('orgName').send_keys(orgname)
	driver.find_element_by_id('EnglishName').send_keys(EnglishName)
	driver.find_element_by_id('address').send_keys(address)
	driver.find_element_by_class_name('ant-upload-btn').click()
	pyperclip.copy('C:\\Users\\DELL\\Pictures\\Feedback\\anxin.png')
	pyperclip.paste()
	driver.find_element_by_class_name('ant-upload-btn').click()
	time.sleep(1)
	#pyperclip.paste()
	pyautogui.typewrite(message=r'C:\Users\DELL\Pictures\Feedback\anxin.png')#,interval=0.5)


def read_xlsx_zuoxi(filename):
	'''读取表格内容 然后创建账号'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	for currt_id_raw in range(23,sheet.max_row+1):
		#sheet.max_row+1
		#账号、姓名、手机号，密码
		user_loginname = sheet.cell(row=currt_id_raw,column=2).value
		user_name = sheet.cell(row=currt_id_raw,column=1).value 
		user_phone = sheet.cell(row=currt_id_raw,column=4).value
		user_password = sheet.cell(row=currt_id_raw,column=3).value
		choice_account(user_loginname,user_name,user_phone,user_password)
def read_xlsx_B(filename):
	'''创建B端账号'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	for currt_id_raw in range(7,sheet.max_row+1):
		#sheet.max_row+1
		user_phone = sheet.cell(row=currt_id_raw,column=5).value
		usr_name = sheet.cell(row=currt_id_raw,column=2).value
		B_choice(user_phone,usr_name,123456)
def location_shubiao():
	x,y=pyautogui.position() #获取当前鼠标的位置
	print(x,y)
	
if __name__ == '__main__':	
	driver=dr()
	current_url()
	login()
	time.sleep(10)
	#setting_button()
	create_jg('test','test','test')
	#read_xlsx_zuoxi(r'C:\Users\DELL\Documents\WeChat Files\sixin2010123\FileStorage\File\2020-09\大案外出智能平台开通账号汇总(1).xlsx')
	#read_xlsx_B(r'D:\project\jbld.xlsx')
	#location_shubiao()
	#driver.close()
