from selenium import webdriver
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import traceback
import os
from selenium.webdriver.chrome.options import Options
import openpyxl
import traceback
import re
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.select import Select
import pyautogui #鼠标点击
import datetime
from selenium.webdriver.common.keys import Keys
from emailSMTP import emailtest #引用邮件
import requests

#启动Chrome;运行cmd的命令
def start_chrome():
	chrome_exe = 'chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile'
	os.popen(chrome_exe)
	time.sleep(3)

def dr():
	chrome_options = Options()
	chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
	chrome_driver = "C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe"
	driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)
	return driver

def login_url():
	GENERATION_URL='https://tpaprod.ikandy.cn/'

	TEST_URL = 'https://tpatest.ikandy.cn/'
	driver.get(GENERATION_URL)
	print('before search======')


#登录
def login_account():
	driver.find_elements_by_class_name('login-input')[0].send_keys('yzc')
	#test = driver.find_element_by_class_name('login-btn')
	test = (By.CLASS_NAME,'login-btn')
	WebDriverWait(driver,20,1).until(EC.presence_of_element_located(test),'没有这个')
	driver.find_elements_by_class_name('login-input')[1].send_keys('123456')
	time.sleep(1)
	driver.find_element_by_class_name('login-btn').click()
	dengl=driver.find_elements_by_class_name('txt-login-form-button')
	time.sleep(2)



def creat_new_case():
	#创建案件
	driver.find_elements_by_class_name('ant-btn')[2].click()
	#
	sfz_id = driver.find_elements_by_class_name('ant-input')[1]
	sfz_id.click()
	sfz_id.send_keys('130635199104240000')
	driver.find_elements_by_class_name('ant-input')[2].click()
	time.sleep(1)
	driver.find_element_by_class_name('ant-calendar-ok-btn').click()
	time.sleep(1)
	driver.



file_name = 'CaseV2-20200308.xlsx'
#读取表格的值
def read_xlsx():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	print(sheet.max_row)
	currt_id = sheet['A2'].value
	sheet['B1'] = 'URL'
	for currt_id_raw in range(2,4):
		#sheet.max_row+1
		produceName = sheet.cell(row=currt_id_raw,column=1).value
		get_url(produceName)
		print(produceName)
		sheet['B'+str(currt_id_raw)] = get_url_txt()
	wb.save(save_filename)

	print(currt_id)
	time.sleep(1)

def get_url(file_url):
	time.sleep(1)
	url_test = ''.join(['https://livecenter-prod.ikandy.cn/api/recordVideos?taskId=',file_url,'&channel=aviva_org'])
	driver.get(url_test)
def get_url_txt():
	time.sleep(1)
	txturl = driver.find_element_by_xpath('/html/body/pre')
	#print(txturl.text)
	regexx=re.compile(r'http.*mp4')
	xd=regexx.findall(txturl.text)
	time.sleep(1)
	return str(xd)


#根据来源填入查询姓名身份证的框
def find_name(test):
	search_name = driver.find_element_by_id('customerName')
	search_name.send_keys(test)
#打印出坐席的名字
def get_name_id():
	time.sleep(2)
	names = driver.find_elements_by_class_name('ant-table-row')
	namess = driver.find_element_by_xpath('//*[@id="root"]/div/section/div[2]/div[2]/section/section/div/div/main/div[2]/div[3]/div/div/div/div/div/div[2]/div/div[1]/table/tbody/tr[1]/td[14]')
	return namess.text



#截图
def current_photo():
	data=time.strftime('%m%d-%H%M',time.localtime(time.time()))
	a=data+'.jpg'
	driver.get_screenshot_as_file(a)


#坐席数
def xlsx_zuoxi():
	r = requests.get(tb_xlsx)
	regexx=re.compile(r'https.*')
	xd=regexx.findall(r.text[:-2])
	dr()
	driver.get(xd)
	print(xd)
	time.sleep(30)
	driver.close()



#读取URL
def read_xlsxx():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	print(sheet.max_row)
	currt_id = sheet['A5'].value
	sheet['C1'] = 'URL'
	for currt_id_raw in range(5,50):
		#sheet.max_row+1
		produceName = sheet.cell(row=currt_id_raw,column=1).value
		sheet['C'+str(currt_id_raw)] = get_requests(produceName)
	wb.save(save_filename)

	print(currt_id)
	time.sleep(1)

def read_xlsx_dzh():
	'''
	file_name = 'CaseV2-20200305.xlsx'
	filename = ''.join(['F:\MyDownloads\\',file_name])
	save_filename = ''.join(['F:\MyDownloads\A',file_name])
	'''
	filename = 'D:\jx\wenjian\周报数据汇总.xlsx'
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	sheet['G3'] = guoren_requests()[0]
	sheet['F3'] = guoren_requests()[1]
	sheet['D6'] = tb_data(tb_URL)
	sheet['D9'] = huanhe_data(hh_URL)
	wb.save(filename)
	time.sleep(1)

#读取时间表
def time_sum():
	filename = r'C:\Users\DELL\Documents\WeChat Files\sixin2010123\FileStorage\20200320report.xlsx'
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	print(sheet.max_row+1)
	for rowNum in range(2,sheet.max_row+1):
		produceName = sheet.cell(row=rowNum,column=1).value
		xin = produceName[3:5]
		sheet.cell(row=rowNum,column=10).value = xin
		xin2 = produceName[-2:]
		sheet.cell(row=rowNum,column=11).value = xin2
	wb.save(filename)


if __name__ == '__main__':

	start_chrome()
	driver=dr()
	login_url()
	login_account()
#	read_xlsx_dzh()
#	time_data()