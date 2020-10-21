import openpyxl #导入操作表格的模板
users=openpyxl.load_workbook('user.xlsx') #打开文件，保存为一个workbook类型的值
users.get_sheet_names() #返回的是一个列表：['sheet1','sheet2']
sheet=users.get_sheet_by_name('test')
print(sheet)
print(sheet.title)

sheetactive=users.get_active_sheet()#取得工作簿的当前活动表

print(sheetactive['A1'].value)	#打印表里A1里面的值
aa=sheetactive['B3']

print('row:'+str(aa.row)+';column:'+str(aa.column)+';coordinate:'+
	aa.coordinate+';value:'+aa.value)

bb=sheetactive.cell(row=4,column=3).value #得到坐标（4.3）的值
print('坐标（4，3）的值：'+str(bb))

#列出A列的值；column=1 固定了列表循环在A列
for i in range(1,4):
	print(i,sheetactive.cell(row=i,column=1).value)
'''
print(type(sheetactive.rows))
for i in sheetactive.rows:
	print(i)
'''

#确定表的大小
max22=sheetactive.max_row
print(str(max22))

#字母和数字之间的转换
from openpyxl.utils import get_column_letter,column_index_from_string
print('数字19代表的字母：'+get_column_letter(19))
print(column_index_from_string('g'))

#从表中取得行和列
print(tuple(sheetactive['A1':'C2']))	#获得A1-C2的数据以行为单位的元组

#print(sheetactive['A1'])
for i in sheetactive['A1':'C3']:
	#print(i)
	for ii in i:	#循环元组的值
		#print(ii)
		print(ii.coordinate,ii.value)
	print('---END---')
cc=list(sheetactive.columns)[3]	#获取D列的列表,用到columns的属性
print(cc)

filessum={}
print('reading rows...')
'''
for row in range(2,sheetactive.max_row+1):
	print(row)
	names=sheetactive['B'+str(row)].value
	phones=sheetactive['C'+str(row)].value
	city=sheetactive['F'+str(row)].value
	filessum.setdefault(names,{}) #setdefault:如果键不存在于字典中，将会添加键并将值设为默认值
	filessum[names].setdefault(phones,{'tracts':0,'city':0})

	filessum[names][phones]['tracts'] += 1
	print(type(city))

	filessum[names][phones]['city']+=int(1) #不明白
import pprint
print('writing results...')
resultsfile=open('username.py','w')
resultsfile.write('allnames='+pprint.pformat(filessum))
resultsfile.close()
'''
#写入Excel
wb=openpyxl.Workbook()
sheetname=wb.get_sheet_names() #获取列表sheet名字
print(sheetname)
sheet=wb.get_active_sheet()
sheet.title='new sheet1'	#修改活动表格sheet的名字
sheetname=wb.get_sheet_names()
print(sheetname)
wb.create_sheet(index=0,title='first sheet') #创建工作表，不加参数默认创建最后一个工作表
wb.create_sheet(index=1,title='test sheet')
sheetname=wb.get_sheet_names()
print(sheetname)
wb.remove_sheet(wb.get_sheet_by_name('test sheet')) #删掉工作表
writesheet=wb.get_active_sheet()
writesheet['A2']='i\'m a beginnner'
names=('yu','zi','chen')
a=0
for i in names:
	a+=1
	print(a)
	writesheet['B'+str(a)]=i
	writesheet['A'+str(a)]='A:'+str(a)
	writesheet['C'+str(a)]='C:'+str(a)


AA=writesheet.cell(row=1,column=1).value
print(AA)

word=[a-z]
a=0
'''
for i in range(5):
	a+=1
	print(writesheet.cell(row=1,column=a).value)
'''
for i in names:
	a+=1
	print('++++')
	print(i)
	print(a)
	print(word[a]+str(a))
	writesheet[word[a]+str(a)]=i

wb.save('example_python.xlsx')


print('开始操作表：')


